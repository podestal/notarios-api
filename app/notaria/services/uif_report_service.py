import io
from datetime import datetime
from typing import Dict, List, Any, Optional

from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class UifReportService:
    """Service for generating UIF reports in HTML and Excel formats."""

    def __init__(self):
        self.header_font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
        self.data_font = Font(name='Arial Narrow', size=10)
        self.header_fill = PatternFill(start_color='254061', end_color='254061', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='376091', end_color='376091', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

    def _transform_record_for_excel(self, record: Dict[str, Any]) -> Dict[str, Any]:
        """Transform a UIF dashboard record into Excel report format."""
        # Get instrument type abbreviation
        tipo_instrumento = {
            1: 'E',  # Escritura
            3: 'T',  # Transferencia
            4: 'G',  # Otros
        }.get(record.get('idtipkar'), 'SIN INICIAL')

        # Format dates
        fecha_escritura = record.get('fechaescritura')
        if isinstance(fecha_escritura, str):
            try:
                fecha_escritura = datetime.strptime(fecha_escritura, '%Y-%m-%d').date()
            except ValueError:
                fecha_escritura = None

        fecha_conclusion = record.get('fechaconclusion')
        if isinstance(fecha_conclusion, str):
            try:
                fecha_conclusion = datetime.strptime(fecha_conclusion, '%Y-%m-%d').date()
            except ValueError:
                fecha_conclusion = None

        # Map data to Excel columns
        return {
            'kardex': record.get('kardex', ''),
            'item_1': record.get('kardex', '')[:8],  # First 8 chars of kardex
            'item_2': str(record.get('idkardex', ''))[:8],  # Numero Registro
            'item_3': 'I',  # Tipo de envio - Initial record
            'item_4': tipo_instrumento,  # Tipo de IPNP
            'item_5': record.get('numescritura', '')[:6],  # Numero del IPNP
            'item_6': fecha_escritura.strftime('%d/%m/%Y') if fecha_escritura else '',  # Fecha del IPNP
            'item_7': '',  # Numero del IPNP que se aclara - empty for initial records
            'item_8': '',  # Fecha del IPNP que se aclara - empty for initial records
            'item_9': '1' if fecha_conclusion else '0',  # Conclusion
            'item_10': fecha_conclusion.strftime('%d/%m/%Y') if fecha_conclusion else '',  # Fecha de firma
            'item_11': '1',  # Modalidad - Single operation
            'item_12': '1',  # Cantidad de operaciones
            'item_13': '',  # Representante
            'item_14': '',  # Persona en cuyo nombre
            'item_15': '',  # Persona a favor
            'item_16': '',  # Persona que representa
            'item_17': '',  # Tipo de representacion
            'item_18': '1',  # Condicion de residencia - default to resident
            'item_19': 'N',  # Tipo de persona - default to Natural
            'item_20': '1',  # Tipo de documento - default to DNI
            'item_21': '',  # Numero de documento
            'item_22': '',  # RUC
            'item_23': '',  # Apellido paterno / Razon social
            'item_24': '',  # Apellido materno
            'item_25': '',  # Nombres
            'item_26': 'PE',  # Pais - default to Peru
            'item_27': '',  # Fecha de nacimiento
            'item_28': '',  # Estado civil
            'item_29': '',  # Codigo de Ocupacion
            'item_30': '',  # Descripcion del objeto social
            'item_31': '',  # Codigo CIIU
            'item_32': '',  # Codigo de Cargo
            'item_33': '',  # Codigo de la Zona Registral
            'item_34': '',  # Numero de la Partida Registral
            'item_35': '',  # Tipo, nombre y numero de la via
            'item_36': '',  # Departamento
            'item_37': '',  # Provincia
            'item_38': '',  # Distrito
            'item_39': '',  # Telefonos
            'item_40': '',  # Participacion del conyuge
            'item_41': '',  # Apellido paterno conyuge
            'item_42': '',  # Apellido materno conyuge
            'item_43': '',  # Nombres conyuge
            'item_44': '',  # Tipo de fondos
            'item_45': record.get('codacto', '')[:3],  # Tipo de operacion - use act code
            'item_46': '',  # Forma de pago
            'item_47': '',  # Oportunidad de pago
            'item_48': '',  # Descripcion de oportunidad de pago
            'item_49': '',  # Origen de los fondos
            'item_50': 'PEN',  # Moneda - default to PEN
            'item_51': '0.00',  # Monto total
            'item_52': '0.00',  # Monto por participante
            'item_53': '0.00',  # Monto relacionado a tipos de fondos
            'item_54': '1.00',  # Tipo de cambio
            'item_55': '',  # Inscripcion registral del bien
            'item_56': '',  # Codigo de la Zona Registral
            'item_57': '',  # Numero de partida registral
        }

    def generate_excel_report(self, data: Dict[str, Any], initial_date: str, final_date: str) -> HttpResponse:
        """Generate Excel report for UIF data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "REGISTRO DE OPERACIONES UIF"

        # Set column widths - these match the PHP script's col width="80" settings
        for col in range(1, 58):  # 57 columns total
            ws.column_dimensions[get_column_letter(col)].width = 12  # Default width

        # Special column widths
        special_widths = {
            23: 40,  # Apellido paterno / Razon social
            30: 40,  # Descripcion del objeto social
            35: 40,  # Tipo, nombre y numero de la via
            49: 40,  # Origen de los fondos
        }
        for col, width in special_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Add headers - exactly like PHP script
        self._add_excel_headers(ws)
        
        # Add data rows
        current_row = 5  # Start after headers
        for record in data.get('lista_kardex_ro', []):
            # Transform record to Excel format
            excel_record = self._transform_record_for_excel(record)
            self._add_excel_data_row(ws, current_row, excel_record)
            current_row += 1

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"UIF_REPORT_{initial_date}_{final_date}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _add_excel_headers(self, ws):
        """Add the complex header structure to Excel worksheet."""
        # Row 1 - Main categories
        headers_row1 = [
            ('', 2),  # Empty cells for first two columns
            ('Datos de identificacion del registro de la operacion', 11),
            ('Participacion y representacion de las personas involucradas en la operacion', 5),
            ('Datos de identificacion de las personas que intervienen en la operacion', 26),
            ('Datos relacionados a la descripcion de la operacion (Acto/Contrato extendido en IPNP)', 14)
        ]
        current_col = 1
        for header, span in headers_row1:
            if span > 1:
                ws.merge_cells(
                    start_row=1, start_column=current_col,
                    end_row=1, end_column=current_col + span - 1
                )
            cell = ws.cell(row=1, column=current_col, value=header)
            self._style_header_cell(cell)
            current_col += span

        # Row 2 - Subcategories
        headers_row2 = [
            ('Numero Registro de la Operacion', 1),
            ('Tipo de envio del RO', 1),
            ('Instrumento Publico Notarial Protocolar (IPNP)', 7),
            ('Modalidad de la operacion', 1),
            ('Cantidad de operaciones individuales que contiene la operacion Multiple', 1),
            ('Roles del Participante', 3),
            ('Representacion', 2),
            ('Condicion de residencia (Declarada en el IPNP)', 1),
            ('Tipo de persona', 1),
            ('Documento de identidad', 2),
            ('Numero de Registro unico de Contribuyente (RUC)', 1),
            ('Nombre completo de la persona', 3),
            ('Pais de nacionalidad', 1),
            ('Fecha de nacimiento', 1),
            ('Estado civil', 1),
            ('Ocupacion, oficio, profesion, actividad economica u objeto social y cargo', 4),
            ('Inscripcion en SUNARP de la Representacion (Personas Juridicas)', 2),
            ('Domicilio y telefonos', 5),
            ('Participacion del conyuge', 1),
            ('Nombre completo del conyuge', 3),
            ('Tipo de fondos, bienes u otros activos con que se realizo la operacion', 1),
            ('Tipo de operacion', 1),
            ('Forma de pago mediante la cual se realizo la operacion', 1),
            ('Oportunidad de pago de la operacion', 1),
            ('Descripcion de la oportunidad de pago (en caso de otros)', 1),
            ('Origen de los fondos, bienes u otros activos involucrados en la operacion', 1),
            ('Moneda en que se realizo la operacion (Codificacion ISO.4217)', 1),
            ('Montos de la operacion', 3),
            ('Tipo de cambio', 1),
            ('Inscripcion en SUNARP del bien materia de la operacion', 3)
        ]
        current_col = 1
        for header, span in headers_row2:
            if span > 1:
                ws.merge_cells(
                    start_row=2, start_column=current_col,
                    end_row=2, end_column=current_col + span - 1
                )
            cell = ws.cell(row=2, column=current_col, value=header)
            self._style_header_cell(cell)
            current_col += span

        # Row 3 - Column names
        headers_row3 = [
            'kardex', 'item: 1', '2', '3', '4', '5', '6', '7', '8', '9', '10',
            '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21',
            '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32',
            '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43',
            '44', '45', '46', '47', '48', '49', '50', '51', '52', '53', '54',
            '55', '56', '57'
        ]
        for col, header in enumerate(headers_row3, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._style_header_cell(cell, fill=self.subheader_fill)

    def _add_excel_data_row(self, ws, row_num: int, record: Dict[str, Any]):
        """Add a data row to Excel worksheet."""
        # Map record data to columns exactly like PHP script
        data = [
            record.get('kardex', ''),  # kardex
            record.get('item_1', '')[:8],  # substr(0,8) like PHP
            record.get('item_2', '')[:8],  # Numero Registro de la Operacion
            record.get('item_3', '')[:1],  # Tipo de envio del RO
            record.get('item_4', '')[:2],  # Tipo de IPNP
            record.get('item_5', '')[:6],  # Numero del IPNP
            record.get('item_6', '')[:8],  # Fecha del IPNP
            record.get('item_7', '')[:6],  # Numero del IPNP que se aclara
            record.get('item_8', '')[:8],  # Fecha del IPNP que se aclara
            record.get('item_9', '')[:1],  # Conclusion
            record.get('item_10', '')[:8],  # Fecha de la firma por participante
            record.get('item_11', '')[:1],  # Modalidad de la operacion
            record.get('item_12', '')[:4],  # Cantidad de operaciones individuales
            record.get('item_13', '')[:1],  # Representante
            record.get('item_14', '')[:1],  # Persona en cuyo nombre se realiza
            record.get('item_15', '')[:1],  # Persona a favor de quien se realiza
            record.get('item_16', '')[:1],  # Persona a la que se representa
            record.get('item_17', '')[:1],  # Tipo de representacion
            record.get('item_18', '')[:1],  # Condicion de residencia
            record.get('item_19', '')[:1],  # Tipo de persona
            record.get('item_20', '')[:1],  # Tipo de documento
            record.get('item_21', '')[:20],  # Numero de documento
            record.get('item_22', '')[:11],  # RUC
            record.get('item_23', '')[:120],  # Apellido paterno / Razon social
            record.get('item_24', '')[:40],  # Apellido materno
            record.get('item_25', '')[:40],  # Nombres
            record.get('item_26', '')[:2],  # Pais de nacionalidad
            record.get('item_27', '')[:8],  # Fecha de nacimiento
            record.get('item_28', '')[:1],  # Estado civil
            record.get('item_29', '')[:3],  # Codigo de Ocupacion
            record.get('item_30', '')[:40],  # Descripcion del objeto social
            record.get('item_31', '')[:4],  # Codigo CIIU
            record.get('item_32', '')[:3],  # Codigo de Cargo
            record.get('item_33', '')[:2],  # Codigo de la Zona Registral
            record.get('item_34', '')[:12],  # Numero de la Partida Registral
            record.get('item_35', '')[:150],  # Tipo, nombre y numero de la via
            record.get('item_36', '')[:2],  # Departamento
            record.get('item_37', '')[:2],  # Provincia
            record.get('item_38', '')[:2],  # Distrito
            record.get('item_39', '')[:40],  # Telefonos
            record.get('item_40', '')[:1],  # Participacion del conyuge
            record.get('item_41', '')[:40],  # Apellido paterno conyuge
            record.get('item_42', '')[:40],  # Apellido materno conyuge
            record.get('item_43', '')[:40],  # Nombres conyuge
            record.get('item_44', '')[:2],  # Tipo de fondos
            record.get('item_45', '')[:3],  # Tipo de operacion
            record.get('item_46', '')[:1],  # Forma de pago
            record.get('item_47', '')[:2],  # Oportunidad de pago
            record.get('item_48', '')[:40],  # Descripcion de oportunidad de pago
            record.get('item_49', '')[:40],  # Origen de los fondos
            record.get('item_50', '')[:3],  # Moneda
            record.get('item_51', '')[:18],  # Monto total
            record.get('item_52', '')[:18],  # Monto por participante
            record.get('item_53', '')[:18],  # Monto relacionado a tipos de fondos
            record.get('item_54', '')[:6],  # Tipo de cambio
            record.get('item_55', '')[:1],  # Inscripcion registral del bien
            record.get('item_56', '')[:2],  # Codigo de la Zona Registral
            record.get('item_57', '')[:12],  # Numero de partida registral
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            self._style_data_cell(cell)
            
            # Apply special number format for specific columns
            if col in [51, 52, 53, 54]:  # Money columns
                cell.number_format = '0.00'
            elif col == 1:  # kardex column
                cell.number_format = '@'  # Text format

    def _style_header_cell(self, cell, fill=None):
        """Apply header styling to a cell."""
        cell.font = self.header_font
        cell.fill = fill or self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border

    def _style_data_cell(self, cell):
        """Apply data styling to a cell."""
        cell.font = self.data_font
        cell.alignment = self.data_alignment
        cell.border = self.border 