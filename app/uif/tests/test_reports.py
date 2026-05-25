import io
from unittest.mock import patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from uif.services.plane_rows import PLANE_BODY_LINE_LENGTH, PLANE_FIELD_PAD, PHP_PLANE_FIELD_WIDTHS
from uif.services.reports import EXCEL_TOTAL_COLUMNS, UifReportService


class ReportTransformTests(SimpleTestCase):
    def test_transform_uses_patrimonial_from_dashboard_record(self):
        service = UifReportService()
        record = {
            "kardex": "K1-2026",
            "idkardex": 1,
            "idtipkar": 1,
            "codacto": "094",
            "uif_code": "010",
            "numescritura": "100",
            "fechaescritura": "2026-04-15",
            "fechaconclusion": "15/04/2026",
            "tipo": "I",
            "tipo_moneda": "SOLES",
            "tipo_cambio": 3.75,
            "patrimonial": 150000,
        }
        row = service._transform_record_for_excel(record, row_index=1)
        self.assertEqual(row["item_51"], "150000.00")
        self.assertEqual(row["item_50"], "PEN")
        self.assertEqual(row["item_45"], "010")

    def test_plane_row_to_excel_row_formats_compact_dates(self):
        service = UifReportService()
        plane_row = {
            "kardex": "K1-2026",
            "item_1": "1",
            "item_6": "20260415",
            "item_10": "20260415",
            "item_27": "15011990",
            "item_9": "C",
        }
        excel_row = service._plane_row_to_excel_row(plane_row)
        self.assertEqual(excel_row["item_6"], "15/04/2026")
        self.assertEqual(excel_row["item_10"], "15/04/2026")
        self.assertEqual(excel_row["item_27"], "15/01/1990")
        self.assertEqual(excel_row["item_9"], "C")

    def test_plane_row_has_php_body_line_length(self):
        service = UifReportService()
        row_values = {f"item_{i}": "" for i in range(1, 58)}
        row_values["item_1"] = "1"
        row_values["item_2"] = "1"
        row_values["item_3"] = "I"
        line = service._format_plane_row(row_values)
        self.assertEqual(len(line), PLANE_BODY_LINE_LENGTH)
        self.assertEqual(PLANE_BODY_LINE_LENGTH, 858)

    def test_plane_row_padding_matches_php_str_pad(self):
        """RoClass::generateFileRo — STR_PAD_LEFT=rjust, STR_PAD_RIGHT=ljust."""
        service = UifReportService()
        row_values = {f"item_{i}": "" for i in range(1, 58)}
        row_values.update(
            {
                "item_1": "1",
                "item_2": "7",
                "item_3": "I",
                "item_4": "E",
                "item_5": "619",
                "item_50": "PEN",
                "item_51": "20000.00",
                "item_52": "0.00",
            }
        )
        line = service._format_plane_row(row_values)
        self.assertEqual(len(line), PLANE_BODY_LINE_LENGTH)
        self.assertTrue(line.startswith("       1       7IE 619   "))
        idx = line.index("PEN")
        self.assertEqual(line[idx : idx + 3], "PEN")
        self.assertEqual(line[idx + 3 : idx + 21], "          20000.00")
        self.assertEqual(line[idx + 21 : idx + 39], "              0.00")

    def test_plane_field_pad_map_matches_php_generate_file_ro(self):
        """Sanity: PLANE_FIELD_PAD letters match PHP str_pad constants per field."""
        php_left = {
            1, 2, 3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 22, 26,
            27, 28, 29, 31, 32, 33, 34, 36, 37, 38, 40, 44, 45, 46, 47, 50, 51, 52,
            53, 54, 55, 56, 57,
        }
        php_right = {4, 5, 21, 23, 24, 25, 30, 35, 39, 41, 42, 43, 48, 49}
        for num in range(1, 58):
            self.assertIn(num, PHP_PLANE_FIELD_WIDTHS)
            if num in php_left:
                self.assertEqual(PLANE_FIELD_PAD[num], "L")
            else:
                self.assertEqual(PLANE_FIELD_PAD[num], "R")

    def test_report_records_prefers_active_list(self):
        service = UifReportService()
        data = {
            "lista_kardex_ro": [{"kardex": "A"}],
            "lista_kardex_report_active": [{"kardex": "A"}, {"kardex": "B"}],
        }
        self.assertEqual(len(service._report_records(data)), 2)

    @patch.object(UifReportService, "_report_records", return_value=[{"kardex": "K1", "tipo": "I", "codacto": "094"}])
    @patch("uif.services.reports.PlaneRowBuilder")
    def test_excel_layout_has_58_columns_and_aligned_headers(
        self, mock_builder_cls, mock_records
    ):
        mock_builder_cls.return_value.build_rows.return_value = [
            {
                "kardex": "K1-2026",
                "item_1": "1",
                "item_2": "1",
                "item_3": "I",
                "item_4": "E",
                "item_5": "1",
                "item_6": "20260423",
                "item_9": "C",
                "item_11": "U",
            }
        ]
        service = UifReportService()
        data = {"summary": {"date_range": {}}, "lista_kardex_ro": []}
        response = service.generate_excel_report(data, "01/04/2026", "30/04/2026")
        self.assertEqual(response.status_code, 200)
        mock_records.assert_called_once()

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.max_column, EXCEL_TOTAL_COLUMNS)
        self.assertEqual(ws.cell(3, 1).value, "kardex")
        self.assertEqual(ws.cell(3, 3).value, "2")
        self.assertEqual(ws.cell(2, 3).value, "Numero Registro de la Operacion")
        self.assertEqual(ws.cell(5, 1).value, "K1-2026")
        self.assertEqual(ws.cell(5, 3).value, "1")
        self.assertEqual(ws.freeze_panes, "C5")
