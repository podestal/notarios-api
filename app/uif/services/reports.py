"""
UIF report generation — Excel and plane (archivo plano) files.
"""

import base64
import io
import logging
import re
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List, Optional

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from notaria.models import Confinotario
from uif.models import RoDataField
from uif.services.plane_rows import (
    PHP_PLANE_FIELD_WIDTHS,
    PLANE_BODY_LINE_LENGTH,
    PLANE_FIELD_PAD,
    PLANE_HEADER_LINE_LENGTH,
    PlaneRowBuilder,
)
from uif.services.report_data import get_uif_report_data, select_report_records

logger = logging.getLogger(__name__)

# Excel shows human-readable dates; plane file keeps compact UIF formats.
EXCEL_DATE_ITEM_NUMBERS = (6, 7, 8, 10, 27)
# kardex + UIF items 1..57
EXCEL_TOTAL_COLUMNS = 58
# PHP substr() limits per item (Excel grid); dates allow full DD/MM/YYYY.
EXCEL_FIELD_LIMITS = {
    1: 30,
    2: 8,
    3: 1,
    4: 2,
    5: 6,
    6: 8,
    7: 6,
    8: 8,
    9: 1,
    10: 8,
    11: 1,
    12: 4,
    13: 1,
    14: 1,
    15: 1,
    16: 1,
    17: 1,
    18: 1,
    19: 1,
    20: 1,
    21: 20,
    22: 11,
    23: 120,
    24: 40,
    25: 40,
    26: 2,
    27: 10,
    28: 1,
    29: 3,
    30: 40,
    31: 4,
    32: 3,
    33: 2,
    34: 12,
    35: 150,
    36: 2,
    37: 2,
    38: 2,
    39: 40,
    40: 1,
    41: 40,
    42: 40,
    43: 40,
    44: 2,
    45: 3,
    46: 1,
    47: 2,
    48: 40,
    49: 40,
    50: 3,
    51: 18,
    52: 18,
    53: 18,
    54: 6,
    55: 1,
    56: 2,
    57: 12,
}


class UifReportService:
    """Excel + plane UIF reports (PHP `_arrObjRo` parity via `reportPolicy=all`)."""

    def __init__(self):
        self.header_font = Font(name="Arial", size=9, color="FFFFFF", bold=True)
        self.data_font = Font(name="Arial Narrow", size=10)
        self.header_fill = PatternFill(start_color="254061", end_color="254061", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="376091", end_color="376091", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        self.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        self._plane_fields: Optional[List[RoDataField]] = None

    def _plane_field_specs(self) -> List[RoDataField]:
        if self._plane_fields is None:
            self._plane_fields = list(
                RoDataField.objects.exclude(number_of_data__isnull=True).order_by("number_of_data")
            )
        return self._plane_fields

    @staticmethod
    def _report_date_range(data: Dict[str, Any]):
        from datetime import datetime

        dr = (data.get("summary") or {}).get("date_range") or {}
        start_iso = dr.get("start_iso")
        end_iso = dr.get("end_iso")
        if not start_iso or not end_iso:
            return None, None
        return (
            datetime.strptime(start_iso, "%Y-%m-%d").date(),
            datetime.strptime(end_iso, "%Y-%m-%d").date(),
        )

    def build_report_data(
        self, initial_date: str, final_date: str, report_policy: str = "all"
    ) -> Dict[str, Any]:
        from uif.services.report_data import parse_report_dates

        start_date, end_date = parse_report_dates(initial_date, final_date)
        return get_uif_report_data(
            start_date, end_date, initial_date, final_date, report_policy=report_policy
        )

    @staticmethod
    def _report_records(data: Dict[str, Any]) -> List[Dict[str, Any]]:
        active = data.get("lista_kardex_report_active")
        if active is not None:
            return list(active)
        policy = data.get("report_policy", "all")
        return select_report_records(data, policy)

    def generate_excel_report(
        self, data: Dict[str, Any], initial_date: str, final_date: str
    ) -> HttpResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = "REGISTRO DE OPERACIONES UIF"

        for col in range(1, EXCEL_TOTAL_COLUMNS + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        for col, width in {6: 12, 8: 12, 10: 12, 27: 12, 23: 40, 30: 40, 35: 40, 49: 40}.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        self._add_excel_headers(ws)
        ws.freeze_panes = "C5"
        date_range = self._report_date_range(data)
        plane_rows = PlaneRowBuilder().build_rows(
            self._report_records(data),
            range_start=date_range[0],
            range_end=date_range[1],
        )
        current_row = 5
        for plane_row in plane_rows:
            excel_row = self._plane_row_to_excel_row(plane_row)
            self._add_excel_data_row(ws, current_row, excel_row)
            current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"UIF_REPORT_{initial_date.replace('/', '-')}_{final_date.replace('/', '-')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def generate_plane_report(
        self, data: Dict[str, Any], initial_date: str, final_date: str
    ) -> HttpResponse:
        notary = Confinotario.objects.first()
        if not notary:
            raise ValueError("No notary configuration found in confinotario")

        final_parts = final_date.split("/")
        year, month, day = final_parts[2], final_parts[1], final_parts[0]
        year_short = year[2:4]
        codigo_oficial = str(notary.codoficial or "").zfill(11)
        codigo_uif = str(notary.coduif or "").zfill(9)
        codigo_uif_ext = str(notary.coduif or "")

        filename = f"04{year_short}{month}{day}501.{codigo_uif_ext}"
        header_core = (
            f"050104     {year}{month}{day}012               {codigo_uif}{codigo_oficial}"
        )
        header_line = header_core.rjust(PLANE_HEADER_LINE_LENGTH)[:PLANE_HEADER_LINE_LENGTH]

        response = HttpResponse(content_type="text/plain; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["X-Filename"] = filename
        response.write(f"{header_line}\r\n")

        date_range = self._report_date_range(data)
        plane_rows = PlaneRowBuilder().build_rows(
            self._report_records(data),
            range_start=date_range[0],
            range_end=date_range[1],
        )
        for row_values in plane_rows:
            line = self._format_plane_row(row_values)
            response.write(f"{line}\r\n")

        return response

    def generate_plane_report_json(
        self, data: Dict[str, Any], initial_date: str, final_date: str
    ) -> Dict[str, Any]:
        """Legacy-compatible JSON wrapper (base64 file content)."""
        http_response = self.generate_plane_report(data, initial_date, final_date)
        filename = http_response.get("X-Filename", "uif_report.txt")
        content_b64 = base64.b64encode(http_response.content).decode("utf-8")
        return {
            "filename": filename,
            "content": content_b64,
            "content_type": "text/plain",
        }

    def _format_plane_row(
        self, row_values: Dict[str, str], specs: Optional[List[RoDataField]] = None
    ) -> str:
        """Fixed-width body line — widths/padding match RoClass::generateFileRo."""
        del specs  # kept for callers; canonical widths come from PHP_PLANE_FIELD_WIDTHS
        parts = []
        for num in range(1, 58):
            width = PHP_PLANE_FIELD_WIDTHS[num]
            raw = str(row_values.get(f"item_{num}", row_values.get(f"field_{num}", "")))
            # PHP STR_PAD_LEFT → rjust; STR_PAD_RIGHT → ljust (RoClass::generateFileRo).
            pad = PLANE_FIELD_PAD.get(num, "L")
            if pad == "L":
                parts.append(raw[:width].rjust(width))
            else:
                parts.append(raw[:width].ljust(width))
        line = "".join(parts)
        if len(line) != PLANE_BODY_LINE_LENGTH:
            logger.warning(
                "Plane row length %s != expected %s",
                len(line),
                PLANE_BODY_LINE_LENGTH,
            )
        return line

    def _transform_record_for_excel(
        self, record: Dict[str, Any], row_index: int = 1
    ) -> Dict[str, str]:
        tipo_instrumento = {1: "E", 3: "T", 4: "G"}.get(record.get("idtipkar"), "SIN INICIAL")

        fecha_escritura = self._format_date(record.get("fechaescritura"))
        fecha_conclusion = self._format_date(record.get("fechaconclusion"))

        patrimonial = record.get("patrimonial")
        if patrimonial is not None:
            patrimonial = self._format_amount(patrimonial)
        else:
            patrimonial = "0.00"

        tipo_cambio = record.get("tipo_cambio")
        if tipo_cambio is not None:
            tipo_cambio = self._format_amount(tipo_cambio, decimals=2)
        else:
            tipo_cambio = "1.00"

        moneda = "USD" if record.get("tipo_moneda") == "DOLARES" else "PEN"
        uif_code = (record.get("uif_code") or record.get("codacto") or "")[:3]

        return {
            "kardex": str(record.get("kardex", "")),
            "item_1": str(row_index),
            "item_2": str(row_index),
            "item_3": str(record.get("tipo", "I"))[:1] or "I",
            "item_4": tipo_instrumento[:2],
            "item_5": str(record.get("numescritura", ""))[:6],
            "item_6": fecha_escritura,
            "item_7": "",
            "item_8": "",
            "item_9": "C" if fecha_conclusion else "N",
            "item_10": fecha_conclusion,
            "item_11": "1",
            "item_12": "1",
            "item_13": "",
            "item_14": "",
            "item_15": "",
            "item_16": "",
            "item_17": "",
            "item_18": "1",
            "item_19": "N",
            "item_20": "1",
            "item_21": "",
            "item_22": "",
            "item_23": "",
            "item_24": "",
            "item_25": "",
            "item_26": "PE",
            "item_27": "",
            "item_28": "",
            "item_29": "",
            "item_30": "",
            "item_31": "",
            "item_32": "",
            "item_33": "",
            "item_34": "",
            "item_35": "",
            "item_36": "",
            "item_37": "",
            "item_38": "",
            "item_39": "",
            "item_40": "",
            "item_41": "",
            "item_42": "",
            "item_43": "",
            "item_44": "",
            "item_45": uif_code,
            "item_46": "",
            "item_47": "",
            "item_48": "",
            "item_49": "",
            "item_50": moneda,
            "item_51": patrimonial,
            "item_52": patrimonial,
            "item_53": "0.00",
            "item_54": tipo_cambio,
            "item_55": "",
            "item_56": "",
            "item_57": "",
        }

    @staticmethod
    def _to_excel_display_date(value) -> str:
        """Plane/PHP compact dates → DD/MM/YYYY for Excel (legacy PHP grid)."""
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        text = str(value).strip()
        if not text:
            return ""
        digits = re.sub(r"\D", "", text)
        if len(digits) == 8:
            try:
                if int(digits[4:6]) <= 12 and int(digits[:4]) > 1900:
                    return datetime.strptime(digits, "%Y%m%d").strftime("%d/%m/%Y")
            except ValueError:
                pass
            try:
                if int(digits[2:4]) <= 12:
                    return datetime.strptime(digits, "%d%m%Y").strftime("%d/%m/%Y")
            except ValueError:
                pass
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text[:10], fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
        return text

    def _plane_row_to_excel_row(self, plane_row: Dict[str, str]) -> Dict[str, str]:
        """Same row structure as plane/_arrObjRo; dates formatted for Excel display."""
        row = dict(plane_row)
        for num in EXCEL_DATE_ITEM_NUMBERS:
            key = f"item_{num}"
            row[key] = self._to_excel_display_date(row.get(key, ""))
        return row

    @staticmethod
    def _format_date(value) -> str:
        if not value:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        text = str(value)
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text[:10], fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
        return text[:10]

    @staticmethod
    def _format_amount(value, decimals: int = 2) -> str:
        try:
            d = Decimal(str(value)).quantize(
                Decimal(10) ** -decimals, rounding=ROUND_HALF_UP
            )
            return format(d, f".{decimals}f")
        except Exception:
            return "0.00"

    def _add_excel_headers(self, ws):
        headers_row1 = [
            ("", 2),
            ("Datos de identificacion del registro de la operacion", 11),
            ("Participacion y representacion de las personas involucradas en la operacion", 5),
            ("Datos de identificacion de las personas que intervienen en la operacion", 26),
            ("Datos relacionados a la descripcion de la operacion (Acto/Contrato extendido en IPNP)", 14),
        ]
        current_col = 1
        for header, span in headers_row1:
            if span > 1:
                ws.merge_cells(
                    start_row=1,
                    start_column=current_col,
                    end_row=1,
                    end_column=current_col + span - 1,
                )
            cell = ws.cell(row=1, column=current_col, value=header)
            self._style_header_cell(cell)
            current_col += span

        headers_row2 = [
            ("", 1),
            ("", 1),
            ("Numero Registro de la Operacion", 1),
            ("Tipo de envio del RO", 1),
            ("Instrumento Publico Notarial Protocolar (IPNP)", 7),
            ("Modalidad de la operacion", 1),
            ("Cantidad de operaciones individuales que contiene la operacion Multiple", 1),
            ("Roles del Participante", 3),
            ("Representacion", 2),
            ("Condicion de residencia (Declarada en el IPNP)", 1),
            ("Tipo de persona", 1),
            ("Documento de identidad", 2),
            ("Numero de Registro unico de Contribuyente (RUC)", 1),
            ("Nombre completo de la persona", 3),
            ("Pais de nacionalidad", 1),
            ("Fecha de nacimiento", 1),
            ("Estado civil", 1),
            ("Ocupacion, oficio, profesion, actividad economica u objeto social y cargo", 4),
            ("Inscripcion en SUNARP de la Representacion (Personas Juridicas)", 2),
            ("Domicilio y telefonos", 5),
            ("Participacion del conyuge", 1),
            ("Nombre completo del conyuge", 3),
            ("Tipo de fondos, bienes u otros activos con que se realizo la operacion", 1),
            ("Tipo de operacion", 1),
            ("Forma de pago mediante la cual se realizo la operacion", 1),
            ("Oportunidad de pago de la operacion", 1),
            ("Descripcion de la oportunidad de pago (en caso de otros)", 1),
            ("Origen de los fondos, bienes u otros activos involucrados en la operacion", 1),
            ("Moneda en que se realizo la operacion (Codificacion ISO.4217)", 1),
            ("Montos de la operacion", 3),
            ("Tipo de cambio", 1),
            ("Inscripcion en SUNARP del bien materia de la operacion", 3),
        ]
        current_col = 1
        for header, span in headers_row2:
            if span > 1:
                ws.merge_cells(
                    start_row=2,
                    start_column=current_col,
                    end_row=2,
                    end_column=current_col + span - 1,
                )
            cell = ws.cell(row=2, column=current_col, value=header)
            self._style_header_cell(cell)
            current_col += span

        headers_row3 = [
            "kardex",
            "item: 1",
            * [str(i) for i in range(2, 58)],
        ]
        for col, header in enumerate(headers_row3, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._style_header_cell(cell, fill=self.subheader_fill)

    def _add_excel_data_row(self, ws, row_num: int, record: Dict[str, str]):
        data = [record.get("kardex", "")]
        for i in range(1, 58):
            lim = EXCEL_FIELD_LIMITS.get(i, 40)
            data.append(str(record.get(f"item_{i}", ""))[:lim])

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            self._style_data_cell(cell)
            if col in (51, 52, 53, 54):
                cell.number_format = "0.00"
            elif col == 1:
                cell.number_format = "@"

    def _style_header_cell(self, cell, fill=None):
        cell.font = self.header_font
        cell.fill = fill or self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border

    def _style_data_cell(self, cell):
        cell.font = self.data_font
        cell.alignment = self.data_alignment
        cell.border = self.border
